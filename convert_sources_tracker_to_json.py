"""
Convert crawl_sources_tracker.xlsx to JSON.

Default:
    .venv/bin/python src/convert_sources_tracker_to_json.py

Custom:
    .venv/bin/python src/convert_sources_tracker_to_json.py \
        --input crawl_sources_tracker.xlsx \
        --sheet Sources \
        --output data/crawl_sources_tracker.json
"""

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).parent
DEFAULT_INPUT = PROJECT_DIR / "crawl_sources_tracker.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "data" / "crawl_sources_tracker.json"
DEFAULT_SHEET = "Sources"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert crawl sources tracker Excel file to JSON."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, type=Path, help="Path to .xlsx file.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet name to convert.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, type=Path, help="Path to output JSON.")
    return parser.parse_args()


def normalize_cell_value(value: Any) -> Any:
    """Convert Excel cell values to JSON-safe values."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def read_sources_tracker(input_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read an Excel sheet into a list of dicts using the first row as headers."""
    workbook = load_workbook(input_path, read_only=True, data_only=True)

    if sheet_name not in workbook.sheetnames:
        available_sheets = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return []

    field_names = [str(header).strip() if header is not None else "" for header in headers]
    if not any(field_names):
        return []

    items = []
    for row in rows:
        if row is None or not any(cell is not None for cell in row):
            continue

        item = {}
        for field_name, value in zip(field_names, row):
            if not field_name:
                continue
            item[field_name] = normalize_cell_value(value)
        items.append(item)

    return items


def main():
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = args.output.expanduser().resolve()

    sources = read_sources_tracker(input_path, args.sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(sources, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Converted {len(sources)} rows from {input_path}")
    print(f"Saved JSON to {output_path}")


if __name__ == "__main__":
    main()
